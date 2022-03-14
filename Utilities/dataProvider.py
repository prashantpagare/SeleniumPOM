import openpyxl


def get_userData(sheetName):
    workbook = openpyxl.load_workbook("..\\excel\\testdata.xlsx")
    sheet = workbook[sheetName]
    rows = sheet.max_row
    cols = sheet.max_column
    mainList = []  # outer list

    for i in range(2, rows + 1):
        dataList = []  # excel data loaded here
        for j in range(1, cols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList
